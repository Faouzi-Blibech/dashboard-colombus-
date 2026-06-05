
import pandas as pd
 
# ── 1. Load ──────────────────────────────────────────────────────────────────
df1 = pd.read_csv(r"dashboard-colombus-/data/fx_clean_data.csv")
df2 = pd.read_csv(r"dashboard-colombus-/data/fx_excel_clean.csv")
 
# ── 2. Drop unwanted columns ─────────────────────────────────────────────────
df1 = df1.drop(columns=["volatility_30d", "risk_level"], errors="ignore")
 
# ── 3. Merge (union) both datasets ───────────────────────────────────────────
df = pd.concat([df1, df2], ignore_index=True)
 
# ── 4. Deduplicate: keep the first occurrence of (date, pair) ────────────────
df["date"] = pd.to_datetime(df["date"])
df = df.drop_duplicates(subset=["date", "pair"]).sort_values(["date", "pair"]).reset_index(drop=True)
 
# ── 5. Pivot: one column per pair (rate) ──────────────────────────────────────
#     Column order is grouped by base currency for readability:
#     EUR/* → GBP/* → USD/* → TND/*
pivot = df.pivot_table(index="date", columns="pair", values="rate", aggfunc="first")
 
BASE_ORDER = ["EUR", "GBP", "USD", "TND"]
 
def pair_sort_key(pair):
    base, quote = pair.split("/")
    b_rank = BASE_ORDER.index(base) if base in BASE_ORDER else 99
    q_rank = BASE_ORDER.index(quote) if quote in BASE_ORDER else 99
    return (b_rank, q_rank)
 
sorted_cols = sorted(pivot.columns, key=pair_sort_key)
pivot = pivot[sorted_cols].reset_index()
 
# ── 6. Also build a "change_pct" pivot for daily_change_pct ──────────────────
chg_pivot = df.pivot_table(index="date", columns="pair", values="daily_change_pct", aggfunc="first")
chg_pivot = chg_pivot[sorted_cols].reset_index()
chg_pivot.columns = ["date"] + [f"{c}_chg%" for c in sorted_cols]
 
# ── 7. Join rate + change pivots ─────────────────────────────────────────────
# Interleave: EUR/GBP | EUR/GBP_chg% | EUR/TND | EUR/TND_chg% | ...
rate_cols  = list(sorted_cols)
chg_cols   = [f"{c}_chg%" for c in sorted_cols]
 
interleaved = ["date"]
for r, c in zip(rate_cols, chg_cols):
    interleaved += [r, c]
 
final = pivot.merge(chg_pivot, on="date")[interleaved]
 
# ── 8. Export ─────────────────────────────────────────────────────────────────
out_path = r"dashboard-colombus-/data/fx_merged.xlsx"
with pd.ExcelWriter(out_path, engine="openpyxl", datetime_format="YYYY-MM-DD") as writer:
    final.to_excel(writer, sheet_name="FX Rates (Wide)", index=False)
    df.to_excel(writer, sheet_name="FX Rates (Long)", index=False)  # keep long form too
 
    # Style the wide sheet
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter
 
    ws = writer.sheets["FX Rates (Wide)"]
 
    # Header styling
    header_fill = PatternFill("solid", start_color="1F3864")
    rate_fill   = PatternFill("solid", start_color="2E4057")
    chg_fill    = PatternFill("solid", start_color="3A5169")
 
    for col_idx, col_name in enumerate(final.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        cell.fill = chg_fill if "_chg%" in str(col_name) else header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
 
    # Number formats
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            col_name = final.columns[cell.column - 1]
            if col_name == "date":
                cell.number_format = "YYYY-MM-DD"
            elif "_chg%" in str(col_name):
                cell.number_format = "0.0000%"
            else:
                cell.number_format = "0.000000"
            cell.alignment = Alignment(horizontal="right")
 
    # Freeze header + date col, auto-fit widths
    ws.freeze_panes = "B2"
    for col_idx, col_name in enumerate(final.columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 14 if col_name != "date" else 12
 
print(f"Done! Saved to {out_path}")
print(f"  Merged rows   : {len(df):,}")
print(f"  Date range    : {df['date'].min().date()} → {df['date'].max().date()}")
print(f"  Pairs (cols)  : {rate_cols}")
 


